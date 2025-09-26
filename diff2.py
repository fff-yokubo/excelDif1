import openpyxl
import sys
import os
from openpyxl import load_workbook

def excel_diff_report(old_file, new_file, output_md="diff_report.md"):
    """
    Excelファイル同士を比較し、セルごとの差分をMarkdown形式で出力する。
    - 長文セルの場合は別枠に「旧値」「新値」を表示
    - 各セルの別枠出力部分の見出しには「差分表へ戻る」リンクを追加
    """

    print(f"[INFO] Start comparison: {old_file} vs {new_file}")

    # ------------------------------------------------------------
    # ファイル存在チェック
    # ------------------------------------------------------------
    if not os.path.exists(old_file):
        print(f"[ERROR] Old file not found: {old_file}")
        return
    if not os.path.exists(new_file):
        print(f"[ERROR] New file not found: {new_file}")
        return

    # ------------------------------------------------------------
    # Excelブックを読み込み（data_only=True で数式を値として取得）
    # ------------------------------------------------------------
    print("[INFO] Loading workbooks...")
    old_wb = load_workbook(old_file, data_only=True)
    new_wb = load_workbook(new_file, data_only=True)

    # ------------------------------------------------------------
    # シート一覧を比較
    # ------------------------------------------------------------
    old_sheets = set(old_wb.sheetnames)
    new_sheets = set(new_wb.sheetnames)

    added_sheets = new_sheets - old_sheets    # 新しく追加されたシート
    removed_sheets = old_sheets - new_sheets  # 削除されたシート
    common_sheets = old_sheets & new_sheets   # 両方に存在するシート

    print(f"[INFO] Sheets found - Added: {added_sheets}, Removed: {removed_sheets}, Common: {common_sheets}")

    # ------------------------------------------------------------
    # Markdownレポート出力開始
    # ------------------------------------------------------------
    with open(output_md, "w", encoding="utf-8") as f:
        # レポートのヘッダ部分
        f.write(f"# Excel差分レポート\n\n")
        f.write(f"- 比較元: `{old_file}`\n")
        f.write(f"- 比較先: `{new_file}`\n\n")

        # 追加されたシートの一覧を出力
        if added_sheets:
            f.write("## 追加されたシート\n\n")
            for s in added_sheets:
                f.write(f"- {s}\n")
            f.write("\n")

        # 削除されたシートの一覧を出力
        if removed_sheets:
            f.write("## 削除されたシート\n\n")
            for s in removed_sheets:
                f.write(f"- {s}\n")
            f.write("\n")

        # --------------------------------------------------------
        # 共通シートごとの差分比較処理
        # --------------------------------------------------------
        for sheet_name in common_sheets:
            print(f"[INFO] Comparing sheet: {sheet_name}")
            f.write(f"## シート: {sheet_name}\n\n")

            old_ws = old_wb[sheet_name]
            new_ws = new_wb[sheet_name]

            # シートの最大行数・列数を取得
            max_row = max(old_ws.max_row, new_ws.max_row)
            max_col = max(old_ws.max_column, new_ws.max_column)

            # 差分結果を保持するテーブル（Markdown用）
            diff_table = []
            # 長文セルの内容を保持するリスト（別枠表示用）
            long_texts = []

            # ----------------------------------------------------
            # 各セルの値を走査し、差分を検出
            # ----------------------------------------------------
            for r in range(1, max_row + 1):
                for c in range(1, max_col + 1):
                    cell = old_ws.cell(r, c).value
                    new_cell = new_ws.cell(r, c).value
                    if cell != new_cell:
                        coord = f"{old_ws.cell(r, c).coordinate}"
                        print(f"[DEBUG] Difference found at {sheet_name} {coord}")

                        # 長文（50文字以上）の場合は別枠出力とする
                        if (cell and len(str(cell)) > 50) or (new_cell and len(str(new_cell)) > 50):
                            # 表にはアンカーリンクだけを置く
                            diff_table.append([
                                coord,
                                f"[旧値はこちら](#{sheet_name}_{coord}_old)",
                                f"[新値はこちら](#{sheet_name}_{coord}_new)"
                            ])
                            # 別枠用にデータを保存
                            long_texts.append((sheet_name, coord, cell, new_cell))
                        else:
                            # 通常の短い文字列はそのまま表に記載
                            diff_table.append([coord, str(cell), str(new_cell)])

            # ----------------------------------------------------
            # 差分表の出力
            # ----------------------------------------------------
            if diff_table:
                # 差分表に戻るためのアンカーを設置
                f.write(f"<a name=\"{sheet_name}_diff_table\"></a>\n\n")
                f.write("| セル | 旧値 | 新値 |\n")
                f.write("|------|------|------|\n")
                for row in diff_table:
                    f.write("| " + " | ".join(row) + " |\n")
                f.write("\n")
            else:
                f.write("変更なし\n\n")

            # ----------------------------------------------------
            # 長文セルの別枠出力
            # ----------------------------------------------------
            for (sheet, coord, old_val, new_val) in long_texts:
                # 見出しに「差分表へ戻る」リンクを追加
                f.write(f"### {sheet} {coord} ([差分表へ戻る](#{sheet}_diff_table))\n")
                if old_val is not None:
                    f.write(f"#### <a name=\"{sheet}_{coord}_old\"></a>旧値\n")
                    f.write("```\n" + str(old_val) + "\n```\n\n")
                if new_val is not None:
                    f.write(f"#### <a name=\"{sheet}_{coord}_new\"></a>新値\n")
                    f.write("```\n" + str(new_val) + "\n```\n\n")

    print(f"[INFO] Diff report generated: {output_md}")


# ------------------------------------------------------------
# メイン処理
# ------------------------------------------------------------
if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python excel_diff.py old.xlsx new.xlsx [output.md]")
    else:
        old_file = sys.argv[1]
        new_file = sys.argv[2]
        output_md = sys.argv[3] if len(sys.argv) > 3 else "diff_report.md"
        excel_diff_report(old_file, new_file, output_md)
