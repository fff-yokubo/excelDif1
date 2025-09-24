#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generate_complex_excel.py
巨大で複雑な架空店舗売上データを生成するスクリプト（完全版）

出力内容:
 - Products (商品マスタ)
 - Stores (店舗マスタ)
 - Employees (従業員サンプル)
 - Sales (トランザクション) : 数式(L: NetSale, M: TotalCost, N: Profit) を埋め込み
 - MonthlySummary : 値として書き出し
 - PivotLike : Category x YearMonth の SUMIFS 数式
 - MonthlyChart : PNG をシートに貼付
 - README : 生成情報

高速化のため、データ生成は pandas で作成し、数式列は文字列 ("=I2*J2*(1-K2)" 形式) を使って
pandas.ExcelWriter(engine="openpyxl") で書き込み、その後 openpyxl で画像挿入を行います。
"""

import argparse
import os
import sys
from datetime import datetime
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

# ----------------------
# ヘルパー / デフォルト
# ----------------------
DEFAULT_ROWS = 20000
NUM_PRODUCTS = 200
NUM_STORES = 50
NUM_EMPLOYEES = 500
START_DATE = pd.to_datetime("2023-01-01")
END_DATE = pd.to_datetime("2025-09-24")  # システム日付に合わせたい場合は datetime.today() 等で差し替え

CURRENCY_FMT = '#,##0.00'

# ----------------------
# メイン生成処理
# ----------------------
def generate_datasets(num_rows=DEFAULT_ROWS, seed=1):
    np.random.seed(seed)
    # --- Stores ---
    regions = ["North", "South", "East", "West", "Central"]
    stores = []
    for i in range(1, NUM_STORES + 1):
        stores.append({
            "StoreID": f"S{i:03}",
            "StoreName": f"Store_{i:03}",
            "Region": np.random.choice(regions),
            "OpenDate": (START_DATE - pd.to_timedelta(np.random.randint(0, 365*5), unit="D")).date(),
            "Manager": f"Manager_{np.random.randint(1, 300):03}"
        })
    df_stores = pd.DataFrame(stores)

    # --- Products ---
    categories = ["Beverages", "Food", "Household", "Electronics", "Clothing", "Toys", "Stationery"]
    products = []
    for i in range(1, NUM_PRODUCTS + 1):
        cat = np.random.choice(categories, p=[0.15,0.25,0.15,0.10,0.15,0.10,0.10])
        products.append({
            "ProductID": f"P{i:04}",
            "ProductName": f"Prod_{i:04}",
            "Category": cat,
            "Cost": float(np.round(np.random.uniform(20, 300), 2)),
            "MSRP": float(np.round(np.random.uniform(50, 1000), 2))
        })
    df_products = pd.DataFrame(products)

    # --- Employees ---
    employees = []
    for i in range(1, NUM_EMPLOYEES + 1):
        employees.append({
            "EmpID": f"E{i:04}",
            "Name": f"Emp_{i:04}",
            "StoreID": df_stores.sample(1).iloc[0]["StoreID"],
            "HireDate": (START_DATE - pd.to_timedelta(np.random.randint(0,365*8), unit="D")).date(),
            "Role": np.random.choice(["Sales","Cashier","Stock","Manager","Support"])
        })
    df_employees = pd.DataFrame(employees)

    # --- Sales (Transactions) ---
    date_range_days = (END_DATE - START_DATE).days + 1
    # ランダムに日付・店舗・商品・数量・単価・割引 を作る
    random_date_offsets = np.random.randint(0, date_range_days, size=num_rows)
    random_dates = [ (START_DATE + pd.Timedelta(days=int(d))).strftime("%Y-%m-%d") for d in random_date_offsets ]
    random_store_idx = np.random.randint(0, NUM_STORES, num_rows)
    random_product_idx = np.random.randint(0, NUM_PRODUCTS, num_rows)
    quantities = np.random.randint(1, 10, num_rows)
    unit_prices = [ df_products.iloc[i]["MSRP"] * (1 + np.random.uniform(-0.3, 0.1)) for i in random_product_idx ]
    discounts = np.round(np.random.choice([0.0, 0.05, 0.10, 0.15, 0.20],
                                         size=num_rows, p=[0.6,0.15,0.12,0.08,0.05]), 2)
    transaction_ids = [ f"T{1000000+i}" for i in range(num_rows) ]

    rows = []
    for i in range(num_rows):
        pid = df_products.iloc[random_product_idx[i]]["ProductID"]
        pname = df_products.iloc[random_product_idx[i]]["ProductName"]
        category = df_products.iloc[random_product_idx[i]]["Category"]
        storeid = df_stores.iloc[random_store_idx[i]]["StoreID"]
        storename = df_stores.iloc[random_store_idx[i]]["StoreName"]
        region = df_stores.iloc[random_store_idx[i]]["Region"]
        qty = int(quantities[i])
        price = float(round(unit_prices[i], 2))
        disc = float(discounts[i])
        # precomputed numeric values (for faster aggregation). Excel will compute formula too.
        net = round(qty * price * (1 - disc), 2)
        cost = float(round(df_products.iloc[random_product_idx[i]]["Cost"] * qty, 2))
        profit = round(net - cost, 2)
        rows.append([
            transaction_ids[i],
            random_dates[i],
            storeid,
            storename,
            region,
            pid,
            pname,
            category,
            qty,
            price,
            disc,
            net,
            cost,
            profit
        ])

    df_sales = pd.DataFrame(rows, columns=[
        "TransactionID","Date","StoreID","StoreName","Region",
        "ProductID","ProductName","Category","Quantity","UnitPrice","Discount","NetSale","TotalCost","Profit"
    ])

    return df_products, df_stores, df_employees, df_sales

def add_formula_columns_to_sales(df_sales):
    """
    df_sales は先に Quantity..Discount までの生データを含むことを想定。
    この関数は NetSale, TotalCost, Profit を Excel 数式文字列として上書きします。
    Excel 列位置（人間のカラム）:
      A TransactionID
      B Date
      C StoreID
      D StoreName
      E Region
      F ProductID
      G ProductName
      H Category
      I Quantity
      J UnitPrice
      K Discount
      L NetSale  <-- formula
      M TotalCost <-- formula (VLOOKUP from Products!A:D col4)
      N Profit <-- formula
    """
    # 行番号1はヘッダー、データは2行目から始まる
    n = len(df_sales)
    # prepare formula columns
    net_formulas = []
    totalcost_formulas = []
    profit_formulas = []
    for i in range(n):
        row_idx = i + 2  # Excel row index
        net_f = f"=I{row_idx}*J{row_idx}*(1-K{row_idx})"
        # VLOOKUP(F{row}, Products!A:D, 4, FALSE) * Quantity
        totalcost_f = f"=VLOOKUP(F{row_idx},Products!A:D,4,FALSE)*I{row_idx}"
        profit_f = f"=L{row_idx}-M{row_idx}"
        net_formulas.append(net_f)
        totalcost_formulas.append(totalcost_f)
        profit_formulas.append(profit_f)

    # Replace numeric precomputed columns with formula strings (Excel will evaluate)
    df = df_sales.copy()
    df.loc[:, "NetSale"] = net_formulas
    df.loc[:, "TotalCost"] = totalcost_formulas
    df.loc[:, "Profit"] = profit_formulas
    return df

def write_to_excel(out_path, df_products, df_stores, df_employees, df_sales_with_formula):
    # まず pandas.ExcelWriter でシートに書き込む（数式文字列もそのまま書き込まれる）
    with pd.ExcelWriter(out_path, engine="openpyxl", mode="w") as writer:
        # Products (write first so that VLOOKUP references exist)
        df_products.to_excel(writer, sheet_name="Products", index=False)
        df_stores.to_excel(writer, sheet_name="Stores", index=False)
        df_employees.to_excel(writer, sheet_name="Employees", index=False)
        df_sales_with_formula.to_excel(writer, sheet_name="Sales", index=False)

        # MonthlySummary を値で書き出す（集計は pandas 側で実施）
        df_sales_vals = df_sales_with_formula.copy()
        # Excel数式が入った列は一時的に評価できないので、集計は precomputed numeric columns from original df needed.
        # For robust operation, compute numeric Net/Cost/Profit from numeric columns present earlier — but since we overwrote
        # those columns with formulas, it's safer to recompute from original numeric info. We'll compute using Quantity, UnitPrice, Discount and Products' Cost.
        # To avoid re-lookup here, we'll recompute NetSale_numeric locally from Quantity/UnitPrice/Discount, and TotalCost_numeric via join to Products.
        # NOTE: df_sales_with_formula still contains UnitPrice, Quantity, Discount columns as numbers.
        df_temp = df_sales_with_formula.copy()
        # net numeric:
        df_temp["NetNumeric"] = df_temp["Quantity"] * df_temp["UnitPrice"] * (1 - df_temp["Discount"])
        # To get TotalCost numeric we need product cost; we'll reload products into a small dict
        # but Products already written to file; we have df_products in memory in caller - so skip here and rely on caller to pass df_products
        # We'll assume caller has df_products accessible via closure; but to be explicit we'll return MonthlySummary in caller.
        # Instead, write a placeholder MonthlySummary here and later we'll overwrite with correct numeric using pandas -> openpyxl reload
        # For simplicity, write an empty MonthlySummary now; we'll fill it after closing writer by re-opening workbook.
        monthly_placeholder = pd.DataFrame(columns=["YearMonth","NetSale","TotalCost","Profit"])
        monthly_placeholder.to_excel(writer, sheet_name="MonthlySummary", index=False)

        # PivotLike (placeholder) — we'll fill with formulas after saving using openpyxl
        pivot_placeholder = pd.DataFrame([["Category"]], columns=["Category"])
        pivot_placeholder.to_excel(writer, sheet_name="PivotLike", index=False)

        # README
        readme = pd.DataFrame([
            ["GeneratedOn", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Description", "Large fictional sales dataset for testing/reporting."],
            ["RowsInSales", str(len(df_sales_with_formula))],
            ["Products", str(len(df_products))],
            ["Stores", str(len(df_stores))],
            ["Employees", str(len(df_employees))],
            ["DateRange", f"{START_DATE.date()} to {END_DATE.date()}"]
        ])
        readme.to_excel(writer, sheet_name="README", index=False, header=False)

    # reopen workbook with openpyxl to modify MonthlySummary (with real numeric aggregates), PivotLike formulas, and add chart image
    from openpyxl import load_workbook
    wb = load_workbook(out_path)
    # read Sales and Products back as DataFrames (for aggregates) using pandas
    # (pandas will read formulas as strings; but original numeric columns Quantity/UnitPrice/Discount are preserved)
    xls = pd.ExcelFile(out_path, engine="openpyxl")
    df_sales_read = pd.read_excel(xls, sheet_name="Sales", engine="openpyxl")
    df_products_read = pd.read_excel(xls, sheet_name="Products", engine="openpyxl")
    # Compute numeric Net/Cost/Profit
    df_sales_read["Date"] = pd.to_datetime(df_sales_read["Date"])
    df_sales_read["NetNumeric"] = df_sales_read["Quantity"] * df_sales_read["UnitPrice"] * (1 - df_sales_read["Discount"])
    # join product cost
    prod_cost_map = df_products_read.set_index("ProductID")["Cost"].to_dict()
    df_sales_read["UnitCost"] = df_sales_read["ProductID"].map(prod_cost_map)
    df_sales_read["TotalCostNumeric"] = df_sales_read["UnitCost"] * df_sales_read["Quantity"]
    df_sales_read["ProfitNumeric"] = df_sales_read["NetNumeric"] - df_sales_read["TotalCostNumeric"]
    df_sales_read["YearMonth"] = df_sales_read["Date"].dt.to_period('M').astype(str)

    monthly = df_sales_read.groupby("YearMonth").agg({
        "NetNumeric":"sum",
        "TotalCostNumeric":"sum",
        "ProfitNumeric":"sum"
    }).reset_index().rename(columns={
        "NetNumeric":"NetSale",
        "TotalCostNumeric":"TotalCost",
        "ProfitNumeric":"Profit"
    }).sort_values("YearMonth")

    # Write MonthlySummary (overwrite)
    ws_m = wb["MonthlySummary"]
    # clear existing
    for row in list(ws_m.rows):
        for cell in row:
            cell.value = None
    # write header + rows
    ws_m.append(["YearMonth","NetSale","TotalCost","Profit"])
    for _, row in monthly.iterrows():
        ws_m.append([row["YearMonth"], float(row["NetSale"]), float(row["TotalCost"]), float(row["Profit"])])

    # PivotLike: Category x YearMonth using SUMIFS formulas
    # We'll build a header row: ["Category", ym1, ym2, ...]
    ws_pv = wb["PivotLike"]
    # clear sheet
    for row in list(ws_pv.rows):
        for cell in row:
            cell.value = None
    yearmonths = monthly["YearMonth"].tolist()
    header = ["Category"] + yearmonths
    ws_pv.append(header)
    categories_unique = df_products_read["Category"].unique().tolist()
    for cat in categories_unique:
        row = [cat]
        for ym in yearmonths:
            # derive start/end date for month
            start_date = pd.to_datetime(ym + "-01").strftime("%Y-%m-%d")
            end_date = (pd.to_datetime(ym + "-01") + pd.offsets.MonthEnd(0)).strftime("%Y-%m-%d")
            formula = f'=SUMIFS(Sales!L:L,Sales!H:H,"{cat}",Sales!B:B,">={start_date}",Sales!B:B,"<={end_date}")'
            row.append(formula)
        ws_pv.append(row)

    # Create chart image from monthly (matplotlib) and insert into MonthlyChart sheet
    import matplotlib
    matplotlib.use('Agg')
    plt.figure(figsize=(10,4))
    plt.plot(monthly["YearMonth"], monthly["NetSale"])
    plt.title("Monthly Net Sales (fictional)")
    plt.xlabel("Year-Month")
    plt.ylabel("Net Sale")
    plt.xticks(rotation=45)
    plt.tight_layout()
    chart_path = os.path.splitext(out_path)[0] + "_monthly_net_sales.png"
    plt.savefig(chart_path)
    plt.close()

    if "MonthlyChart" not in wb.sheetnames:
        wb.create_sheet("MonthlyChart")
    ws_chart = wb["MonthlyChart"]
    img = XLImage(chart_path)
    img.anchor = "A1"
    ws_chart.add_image(img)

    # Optionally apply some number formats to MonthlySummary and PivotLike result cells
    # MonthlySummary: columns B-D are numeric
    for row in ws_m.iter_rows(min_row=2, min_col=2, max_col=4):
        for cell in row:
            cell.number_format = CURRENCY_FMT

    # Save workbook
    wb.save(out_path)
    # clean up chart image file
    try:
        os.remove(chart_path)
    except Exception:
        pass

def main():
    parser = argparse.ArgumentParser(description="Generate a large complex Excel file with fictional store sales data.")
    parser.add_argument("--rows", type=int, default=DEFAULT_ROWS, help="Number of sales rows to generate (default: 20000).")
    parser.add_argument("--out", type=str, default="complex_huge_sample.xlsx", help="Output xlsx filename (default: complex_huge_sample.xlsx).")
    parser.add_argument("--seed", type=int, default=1, help="Random seed (default: 1).")
    args = parser.parse_args()

    rows = args.rows
    out_path = args.out
    seed = args.seed

    print(f"Generating datasets (rows={rows}, seed={seed}) ...")
    df_products, df_stores, df_employees, df_sales = generate_datasets(num_rows=rows, seed=seed)
    print("Adding Excel formula columns to Sales (NetSale, TotalCost, Profit) ...")
    df_sales_with_formula = add_formula_columns_to_sales(df_sales)

    print(f"Writing to Excel: {out_path} ... (this may take a while for large rows)")
    write_to_excel(out_path, df_products, df_stores, df_employees, df_sales_with_formula)
    print("Done.")
    print(f"Output file: {os.path.abspath(out_path)} (rows in Sales = {len(df_sales_with_formula)})")

if __name__ == "__main__":
    main()
